import base64
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
import pytz

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

from openpyxl import load_workbook

SCOPES = ["https://www.googleapis.com/auth/gmail.modify"]

SENDER = "miriam.martinescu@ligaac.ro"
CC_LIST = [
    "stefan.covaci@ligaac.ro",
    "robert.gaube@ligaac.ro"
]

SCHEDULE_TIME = datetime(
    2025, 12, 26, 8, 0, 0,
    tzinfo=pytz.timezone("Europe/Bucharest")
)

EXCEL_FILE = "output.xlsx"


def gmail_service():
    creds = None
    try:
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    except:
        pass

    if not creds or not creds.valid:
        flow = InstalledAppFlow.from_client_secrets_file(
            "credentials.json", SCOPES
        )
        creds = flow.run_local_server(port=0)
        with open("token.json", "w") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def create_message(to, subject, body):
    msg = MIMEMultipart()
    msg["From"] = SENDER
    msg["To"] = to
    msg["Cc"] = ", ".join(CC_LIST)
    msg["Subject"] = subject

    msg.attach(MIMEText(body, "plain", "utf-8"))

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    return {"raw": raw}


def main():
    service = gmail_service()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    send_at = int(SCHEDULE_TIME.timestamp())

    for row in ws.iter_rows(min_row=2, values_only=True):
        name, phone, email, link = row

        if not email or not link:
            continue

        body = f"""Ho ho ho! 🎅❄️

După ce a rătăcit printre fulgi, cadouri și căni de vin fiert, Moșul a ajuns în sfârșit cu scrisoarea ta de Crăciun 🎄✨

👉 Scrisoarea ta te așteaptă aici:
{link}

Sperăm să-ți aducă un zâmbet și un pic de magie de sărbători ✨

Pupicei! 💙
"""

        message = create_message(
            to=email,
            subject="🎄 Scrisoarea ta de Crăciun a sosit!",
            body=body
        )

        draft = service.users().drafts().create(
            userId="me",
            body={"message": message}
        ).execute()

        service.users().drafts().send(
            userId="me",
            body={
                "id": draft["id"],
                "sendAt": send_at
            }
        ).execute()

        print(f"✔ Programat pentru {email}")

    print("\n🎉 TOATE EMAILURILE AU FOST PROGRAMATE CU SUCCES!")


if __name__ == "__main__":
    main()
