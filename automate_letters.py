import warnings
warnings.filterwarnings("ignore", category=UserWarning)

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
import time
import os

URL = "https://oscrisoare.ro/compune.php"
IMAGE_PATH = os.path.abspath("img.jpeg")
INPUT_EXCEL = "input.xlsx"
OUTPUT_EXCEL = "output.xlsx"

def wait_click(driver, wait, by, value):
    el = wait.until(EC.element_to_be_clickable((by, value)))
    driver.execute_script("arguments[0].click();", el)

def fill_text(driver, wait, by, value, text):
    el = wait.until(EC.presence_of_element_located((by, value)))
    driver.execute_script("arguments[0].value = '';", el)
    el.send_keys(text)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
wait = WebDriverWait(driver, 30)

wb = load_workbook(INPUT_EXCEL)
ws = wb.active

out_wb = Workbook()
out_ws = out_wb.active
out_ws.append(["Name", "Letter Link"])

for row in ws.iter_rows(values_only=True):
    name = row[0]
    message = row[2]

    if not name or not message:
        continue

    name = str(name).strip()
    message = str(message).strip()

    paragraphs = [p.strip() for p in message.split("\n\n") if p.strip()]
    if len(paragraphs) < 2:
        continue

    title = paragraphs[0]
    first = paragraphs[1]
    second = paragraphs[2] if len(paragraphs) > 2 else ""
    final = "\n\n".join(paragraphs[3:]) if len(paragraphs) > 3 else ""

    driver.get(URL)

    wait_click(driver, wait, By.NAME, "NEW")
    time.sleep(1)

    fill_text(driver, wait, By.ID, "page_title", "Scrisoare sărbători")
    fill_text(driver, wait, By.ID, "title", title)
    fill_text(driver, wait, By.ID, "text1", first)
    fill_text(driver, wait, By.ID, "text2", second)
    fill_text(driver, wait, By.ID, "text3", final)

    wait_click(driver, wait, By.CSS_SELECTOR, "a.fonturi")
    time.sleep(0.5)

    Select(wait.until(EC.presence_of_element_located((By.ID, "font_title")))).select_by_value("Gilda+Display")
    Select(driver.find_element(By.ID, "font_text1")).select_by_value("Gilda+Display")
    Select(driver.find_element(By.ID, "font_text2")).select_by_value("Gilda+Display")
    Select(driver.find_element(By.ID, "font_text3")).select_by_value("Gilda+Display")

    wait_click(driver, wait, By.CSS_SELECTOR, "a.imagini")
    time.sleep(0.5)

    driver.find_element(By.ID, "file_upload").send_keys(IMAGE_PATH)
    time.sleep(0.5)

    wait_click(driver, wait, By.CSS_SELECTOR, "input[name='submit'][value='Upload']")
    time.sleep(1.5)

    wait_click(driver, wait, By.ID, "save")
    time.sleep(1)

    link_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input.link")))
    link = link_input.get_attribute("value")

    out_ws.append([name, link])

out_wb.save(OUTPUT_EXCEL)
driver.quit()
